# DJANGO IMPORTS
from django.views.generic import View
from django.http import HttpResponse

# THIRD PARTY IMPORTS
import openpyxl

# MODELS IMPORTS
from .models import Student


class StudentView(View):
    def get(self, request):
        students = Student.objects.all()
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="students.xlsx"' # noqa

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Students'

        ws['A1'] = 'Name'
        ws['B1'] = 'Age'
        ws['C1'] = 'Grade'

        row = 2
        for student in students:
            ws.cell(row=row, column=1, value=student.name)
            ws.cell(row=row, column=2, value=student.age)
            ws.cell(row=row, column=3, value=student.grade)
            row += 1

        wb.save(response)
        return response
